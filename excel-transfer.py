from openpyxl import Workbook, workbook, load_workbook
from pathlib import Path

data_folder = Path("C:/Users/hasan/Desktop/Test/NAS")#change path to connected network path
#data_folder = Path("/volume1/home/admin/ZEUS", "rb")
file_to_open = data_folder / "HSvideos_and_LLS_11May2020_MCF.xlsx"

workbook = load_workbook(file_to_open, data_only=True)

sheet = workbook['HS_data']
#print(sheet)


reference = []
for ref in sheet.iter_cols(min_col=1, max_col=1, min_row=2, values_only=True):
    reference.append(ref)
#print(ref[6])

strokeChannelNum = []
for scn in sheet.iter_cols(min_col=2, max_col=2, min_row=2, values_only=True):
    strokeChannelNum.append(scn)
#print(strokeChannelNum[6])

timeAbs = []
for ta in sheet.iter_cols(min_col=3, max_col=3, min_row=2, values_only=True):
    timeAbs.append(ta)
#print(timeAbs[6])

thunderstormDay = []
for td in sheet.iter_cols(min_col=4, max_col=4, min_row=2, values_only=True):
    thunderstormDay.append(td)
#print(thunderstormDay[6])

year = []
for y in sheet.iter_cols(min_col=5, max_col=5, min_row=2, values_only=True):
    year.append(y)

month = []
for m in sheet.iter_cols(min_col=6, max_col=6, min_row=2, values_only=True):
    month.append(m)

day = []
for d in sheet.iter_cols(min_col=7, max_col=7, min_row=2, values_only=True):
    day.append(d)

hour = []
for h in sheet.iter_cols(min_col=8, max_col=8, min_row=2, values_only=True):
    hour.append(h)

minute = []
for min in sheet.iter_cols(min_col=9, max_col=9, min_row=2, values_only=True):
    minute.append(min)

second = []
for s in sheet.iter_cols(min_col=10, max_col=10, min_row=2, values_only=True):
    second.append(s)

millisecond = []
for milli in sheet.iter_cols(min_col=11, max_col=11, min_row=2, values_only=True):
    millisecond.append(milli)

process = []
for p in sheet.iter_cols(min_col=12, max_col=12, min_row=2, values_only=True):
    process.append(p)

strikePoint = []
for sp in sheet.iter_cols(min_col=13, max_col=13, min_row=2, values_only=True):
    strikePoint.append(sp)

polarity = []
for pol in sheet.iter_cols(min_col=14, max_col=14, min_row=2, values_only=True):
    polarity.append(pol)

visibility = []
for vis in sheet.iter_cols(min_col=16, max_col=16, min_row=2, values_only=True):
    visibility.append(vis)

duration = []
for dur in sheet.iter_cols(min_col=17, max_col=17, min_row=2, values_only=True):
    duration.append(dur)